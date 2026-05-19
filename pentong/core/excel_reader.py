"""엑셀 파일 읽기 — openpyxl 기반.

시트 목록 조회, 전체/범위 읽기, 헤더 감지 등.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def get_workbook_info(filepath: Path) -> dict:
    """엑셀 파일의 기본 정보를 반환한다.

    Returns:
        {
            "filepath": str,
            "sheet_names": [str, ...],
            "sheet_count": int,
            "sheets": [
                {"name": str, "rows": int, "cols": int, "header": [str, ...]},
                ...
            ]
        }
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets_info = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0

        # 첫 행을 헤더로
        header = []
        if rows > 0:
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                header.append(str(cell) if cell is not None else "")

        sheets_info.append({
            "name": name,
            "rows": rows,
            "cols": cols,
            "header": header,
        })

    wb.close()
    return {
        "filepath": str(filepath),
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets_info,
    }


def read_sheet(
    filepath: Path,
    sheet_name: str | None = None,
    header_row: int = 1,
    as_dicts: bool = False,
) -> list[list] | list[dict]:
    """시트의 전체 데이터를 읽는다.

    Args:
        filepath: xlsx 파일 경로.
        sheet_name: 시트 이름. None이면 활성 시트.
        header_row: 헤더 행 번호 (1부터). as_dicts=True일 때 사용.
        as_dicts: True면 [{컬럼명: 값}, ...] 형태로 반환.

    Returns:
        2차원 리스트 또는 딕셔너리 리스트.
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    if as_dicts and len(rows) > header_row - 1:
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[header_row - 1])]
        data_rows = rows[header_row:]
        return [
            {headers[i]: val for i, val in enumerate(row) if i < len(headers)}
            for row in data_rows
        ]

    return [list(row) for row in rows]


def read_range(
    filepath: Path,
    sheet_name: str | None = None,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> list[list]:
    """시트의 특정 범위를 읽는다.

    Args:
        filepath: xlsx 파일 경로.
        sheet_name: 시트 이름. None이면 활성 시트.
        min_row, max_row: 행 범위 (1부터).
        min_col, max_col: 열 범위 (1부터).

    Returns:
        2차원 리스트.
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ))
    wb.close()

    return [list(row) for row in rows]


def search_cells(
    filepath: Path,
    keyword: str,
    sheet_name: str | None = None,
    ignore_case: bool = True,
) -> list[dict]:
    """엑셀에서 특정 텍스트를 포함하는 셀을 검색한다.

    Returns:
        [
            {"sheet": str, "row": int, "col": int, "value": str},
            ...
        ]
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    results: list[dict] = []

    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_str = str(cell.value)
                match = False
                if ignore_case:
                    match = keyword.lower() in cell_str.lower()
                else:
                    match = keyword in cell_str

                if match:
                    results.append({
                        "sheet": sname,
                        "row": cell.row,
                        "col": cell.column,
                        "value": cell_str,
                    })

    wb.close()
    return results


def get_column_values(
    filepath: Path,
    column: int,
    sheet_name: str | None = None,
    skip_header: bool = True,
    unique: bool = False,
) -> list:
    """특정 열의 모든 값을 반환한다.

    Args:
        filepath: xlsx 파일 경로.
        column: 열 번호 (1부터).
        sheet_name: 시트 이름.
        skip_header: 첫 행(헤더) 건너뛰기.
        unique: 중복 제거.

    Returns:
        값 리스트.
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_row = 2 if skip_header else 1
    values = []
    for row in ws.iter_rows(min_row=start_row, min_col=column, max_col=column, values_only=True):
        values.append(row[0])

    wb.close()

    if unique:
        seen = set()
        result = []
        for v in values:
            if v not in seen:
                seen.add(v)
                result.append(v)
        return result

    return values
