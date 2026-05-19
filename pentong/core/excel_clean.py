"""엑셀 공백 행 제거 로직.

공백 행 판정: 행의 모든 셀이 None이거나 빈 문자열(strip 후)인 행.
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def count_blank_rows(input_path: Path) -> int:
    """파일 내 공백 행 수를 반환한다.

    Args:
        input_path: 검사할 xlsx 파일 경로.

    Returns:
        공백 행 수.

    Raises:
        FileNotFoundError: 파일이 존재하지 않을 때.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_path}")

    wb = openpyxl.load_workbook(input_path)
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            if _is_blank_row(row):
                count += 1
    wb.close()
    return count


def remove_blank_rows(
    input_path: Path,
    output_path: Path,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> int:
    """공백 행을 제거한 새 파일을 저장한다.

    모든 시트의 공백 행을 제거하며 나머지 서식은 보존한다.

    Args:
        input_path: 원본 xlsx 파일 경로.
        output_path: 저장할 xlsx 파일 경로 (원본과 같으면 덮어쓰기).
        progress_cb: (현재_시트, 전체_시트, 시트명) 형태의 콜백. 생략 가능.

    Returns:
        제거된 공백 행의 총 수.

    Raises:
        FileNotFoundError: 입력 파일이 존재하지 않을 때.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_path}")

    src_wb = openpyxl.load_workbook(input_path)
    sheets = src_wb.sheetnames
    total = len(sheets)
    removed = 0

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    for idx, sheet_name in enumerate(sheets):
        if progress_cb:
            progress_cb(idx + 1, total, sheet_name)

        src_ws: Worksheet = src_wb[sheet_name]
        dest_ws = out_wb.create_sheet(title=sheet_name)

        data_rows = list(src_ws.iter_rows())
        dest_row = 1
        for row in data_rows:
            if _is_blank_row(row):
                removed += 1
                continue
            for cell in row:
                dest_cell = dest_ws.cell(row=dest_row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dest_cell.font = cell.font.copy()
                    dest_cell.border = cell.border.copy()
                    dest_cell.fill = cell.fill.copy()
                    dest_cell.number_format = cell.number_format
                    dest_cell.alignment = cell.alignment.copy()
            dest_row += 1

        # 열 너비 보존
        for col_idx, col_dim in src_ws.column_dimensions.items():
            dest_ws.column_dimensions[col_idx].width = col_dim.width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)
    src_wb.close()
    return removed


# ---------------------------------------------------------------------------
# 내부 유틸
# ---------------------------------------------------------------------------

def _is_blank_row(row: tuple) -> bool:
    """행의 모든 셀이 비어있으면 True."""
    for cell in row:
        val = cell.value
        if val is None:
            continue
        if isinstance(val, str) and val.strip() == "":
            continue
        return False
    return True
