"""엑셀 파일 분할 로직.

세 가지 분할 방식 지원:
  - 'sheet': 각 시트를 개별 파일로 저장
  - 'rows': N행 단위로 분할
  - 'column': 특정 열의 고유값 기준으로 분할
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Literal

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


SplitMode = Literal["sheet", "rows", "column"]


def split_excel_file(
    input_path: Path,
    output_dir: Path,
    mode: SplitMode = "sheet",
    rows_per_file: int = 1000,
    column_index: int = 1,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> list[Path]:
    """하나의 xlsx 파일을 여러 파일로 분할한다.

    Args:
        input_path: 분할할 xlsx 파일 경로.
        output_dir: 결과 파일을 저장할 폴더.
        mode: 'sheet' — 시트별 분할; 'rows' — N행 단위; 'column' — 열값 기준.
        rows_per_file: mode='rows'일 때 파일당 최대 행 수 (헤더 제외).
        column_index: mode='column'일 때 기준 열 번호 (1부터 시작).
        progress_cb: (현재_인덱스, 전체_수, 설명) 형태의 콜백. 생략 가능.

    Returns:
        생성된 파일 경로 목록.

    Raises:
        ValueError: 잘못된 파라미터.
        FileNotFoundError: 입력 파일이 존재하지 않을 때.
    """
    if not input_path.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {input_path}")
    if mode == "rows" and rows_per_file < 1:
        raise ValueError("rows_per_file은 1 이상이어야 합니다.")
    if mode == "column" and column_index < 1:
        raise ValueError("column_index는 1 이상이어야 합니다.")

    output_dir.mkdir(parents=True, exist_ok=True)

    if mode == "sheet":
        return _split_by_sheet(input_path, output_dir, progress_cb)
    elif mode == "rows":
        return _split_by_rows(input_path, output_dir, rows_per_file, progress_cb)
    else:
        return _split_by_column(input_path, output_dir, column_index, progress_cb)


# ---------------------------------------------------------------------------
# 내부 구현
# ---------------------------------------------------------------------------

def _split_by_sheet(
    input_path: Path,
    output_dir: Path,
    progress_cb: Callable[[int, int, str], None] | None,
) -> list[Path]:
    """각 시트를 개별 xlsx 파일로 저장한다."""
    src_wb = openpyxl.load_workbook(input_path)
    sheets = src_wb.sheetnames
    total = len(sheets)
    stem = input_path.stem
    results: list[Path] = []

    for idx, sheet_name in enumerate(sheets):
        if progress_cb:
            progress_cb(idx + 1, total, sheet_name)

        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        dest_ws = out_wb.create_sheet(title=sheet_name)
        _copy_sheet(src_wb[sheet_name], dest_ws)

        safe_name = _safe_filename(sheet_name)
        out_path = output_dir / f"{stem}_{safe_name}.xlsx"
        out_wb.save(out_path)
        results.append(out_path)

    src_wb.close()
    return results


def _split_by_rows(
    input_path: Path,
    output_dir: Path,
    rows_per_file: int,
    progress_cb: Callable[[int, int, str], None] | None,
) -> list[Path]:
    """활성 시트를 N행 단위로 분할한다. 헤더는 모든 파일에 포함된다."""
    src_wb = openpyxl.load_workbook(input_path)
    src_ws = src_wb.active
    stem = input_path.stem

    all_rows = list(src_ws.iter_rows(values_only=False))
    if not all_rows:
        src_wb.close()
        return []

    header_rows = all_rows[:1]
    data_rows = all_rows[1:]

    chunks = [data_rows[i:i + rows_per_file] for i in range(0, max(len(data_rows), 1), rows_per_file)]
    if not chunks:
        chunks = [[]]
    total = len(chunks)
    results: list[Path] = []

    for idx, chunk in enumerate(chunks):
        label = f"{idx + 1:03d}"
        if progress_cb:
            progress_cb(idx + 1, total, label)

        out_wb = Workbook()
        dest_ws = out_wb.active
        dest_ws.title = src_ws.title

        for row in header_rows:
            dest_ws.append([cell.value for cell in row])
        for row in chunk:
            dest_ws.append([cell.value for cell in row])

        _copy_col_widths(src_ws, dest_ws)

        out_path = output_dir / f"{stem}_{label}.xlsx"
        out_wb.save(out_path)
        results.append(out_path)

    src_wb.close()
    return results


def _split_by_column(
    input_path: Path,
    output_dir: Path,
    column_index: int,
    progress_cb: Callable[[int, int, str], None] | None,
) -> list[Path]:
    """활성 시트에서 특정 열의 고유값 기준으로 분할한다. 헤더는 모든 파일에 포함된다."""
    src_wb = openpyxl.load_workbook(input_path)
    src_ws = src_wb.active
    stem = input_path.stem

    all_rows = list(src_ws.iter_rows(values_only=False))
    if not all_rows:
        src_wb.close()
        return []

    header_rows = all_rows[:1]
    data_rows = all_rows[1:]

    # 열값 기준 그룹화 (순서 유지)
    groups: dict[str, list] = {}
    for row in data_rows:
        if column_index <= len(row):
            key = str(row[column_index - 1].value) if row[column_index - 1].value is not None else "__빈값__"
        else:
            key = "__빈값__"
        groups.setdefault(key, []).append(row)

    keys = list(groups.keys())
    total = len(keys)
    results: list[Path] = []

    for idx, key in enumerate(keys):
        if progress_cb:
            progress_cb(idx + 1, total, key)

        out_wb = Workbook()
        dest_ws = out_wb.active
        dest_ws.title = src_ws.title

        for row in header_rows:
            dest_ws.append([cell.value for cell in row])
        for row in groups[key]:
            dest_ws.append([cell.value for cell in row])

        _copy_col_widths(src_ws, dest_ws)

        safe_key = _safe_filename(key)
        out_path = output_dir / f"{stem}_{safe_key}.xlsx"
        out_wb.save(out_path)
        results.append(out_path)

    src_wb.close()
    return results


def _copy_sheet(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    """시트 내용과 서식을 복사한다."""
    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = cell.font.copy()
                dest_cell.border = cell.border.copy()
                dest_cell.fill = cell.fill.copy()
                dest_cell.number_format = cell.number_format
                dest_cell.alignment = cell.alignment.copy()

    _copy_col_widths(src_ws, dest_ws)

    for row_idx, row_dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row_idx].height = row_dim.height


def _copy_col_widths(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    for col_idx, col_dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_idx].width = col_dim.width


def _safe_filename(name: str) -> str:
    """파일명으로 사용할 수 없는 문자를 제거한다."""
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name[:50]
