"""엑셀 파일 취합 로직.

두 가지 취합 방식 지원:
  - 'sheet': 각 파일을 별도 시트로 추가
  - 'row': 모든 파일의 행을 단일 시트에 연결
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Literal

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


MergeMode = Literal["sheet", "row"]


def merge_excel_files(
    input_paths: list[Path],
    output_path: Path,
    mode: MergeMode = "sheet",
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> None:
    """여러 xlsx 파일을 하나의 파일로 취합한다.

    Args:
        input_paths: 취합할 xlsx 파일 경로 목록.
        output_path: 저장할 출력 파일 경로.
        mode: 'sheet' — 파일마다 시트 생성; 'row' — 모든 행을 한 시트에 연결.
        progress_cb: (현재_인덱스, 전체_수, 현재_파일명) 형태의 콜백. 생략 가능.

    Raises:
        ValueError: input_paths가 비어 있을 때.
        FileNotFoundError: 지정한 파일이 존재하지 않을 때.
    """
    if not input_paths:
        raise ValueError("취합할 파일을 하나 이상 선택하세요.")

    for p in input_paths:
        if not p.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {p}")

    if mode == "sheet":
        _merge_as_sheets(input_paths, output_path, progress_cb)
    else:
        _merge_as_rows(input_paths, output_path, progress_cb)


# ---------------------------------------------------------------------------
# 내부 구현
# ---------------------------------------------------------------------------

def _merge_as_sheets(
    input_paths: list[Path],
    output_path: Path,
    progress_cb: Callable[[int, int, str], None] | None,
) -> None:
    """각 파일의 활성 시트를 별도 시트로 추가한다. 서식 포함."""
    out_wb = Workbook()
    out_wb.remove(out_wb.active)  # 기본 빈 시트 제거

    total = len(input_paths)
    for idx, src_path in enumerate(input_paths):
        if progress_cb:
            progress_cb(idx + 1, total, src_path.name)

        src_wb = openpyxl.load_workbook(src_path)
        sheet_name = _unique_sheet_name(out_wb, src_path.stem)

        dest_ws = out_wb.create_sheet(title=sheet_name)
        src_ws = src_wb.active

        _copy_sheet(src_ws, dest_ws)
        src_wb.close()

    out_wb.save(output_path)


def _merge_as_rows(
    input_paths: list[Path],
    output_path: Path,
    progress_cb: Callable[[int, int, str], None] | None,
) -> None:
    """모든 파일의 행을 단일 시트에 이어 붙인다. 첫 파일의 헤더만 유지."""
    out_wb = Workbook()
    dest_ws = out_wb.active
    dest_ws.title = "취합결과"

    first_file = True
    total = len(input_paths)

    for idx, src_path in enumerate(input_paths):
        if progress_cb:
            progress_cb(idx + 1, total, src_path.name)

        src_wb = openpyxl.load_workbook(src_path)
        src_ws = src_wb.active

        start_row = 1 if first_file else 2  # 두 번째 파일부터 헤더(첫 행) 제외
        for row in src_ws.iter_rows(min_row=start_row, values_only=False):
            new_row_values = [cell.value for cell in row]
            dest_ws.append(new_row_values)

        first_file = False
        src_wb.close()

    out_wb.save(output_path)


def _copy_sheet(src_ws, dest_ws) -> None:
    """시트 내용과 컬럼 너비를 복사한다."""
    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest_cell.font = cell.font.copy()
                dest_cell.border = cell.border.copy()
                dest_cell.fill = cell.fill.copy()
                dest_cell.number_format = cell.number_format
                dest_cell.alignment = cell.alignment.copy()

    # 컬럼 너비 복사
    for col_idx, col_dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_idx].width = col_dim.width

    # 행 높이 복사
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row_idx].height = row_dim.height


def _unique_sheet_name(wb: Workbook, base: str) -> str:
    """wb 내에서 중복되지 않는 시트 이름을 반환한다. 최대 31자 제한."""
    existing = {ws.title for ws in wb.worksheets}
    name = base[:31]
    if name not in existing:
        return name
    for i in range(2, 10000):
        candidate = f"{base[:28]}_{i}"
        if candidate not in existing:
            return candidate
    raise RuntimeError("시트 이름 생성 실패")
