"""엑셀 양식 채우기 — openpyxl 기반.

양식 엑셀 파일에서 플레이스홀더를 탐지하고,
사본을 만든 뒤 셀 값을 채워넣는다.

플레이스홀더 규칙:
  - {{필드명}} 형식 (권장)
  - {필드명} 형식
  - OOO, ○○○ 등 반복 문자
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Callable

import openpyxl


# 플레이스홀더 탐지 패턴들
PLACEHOLDER_PATTERNS = [
    r"\{\{(.+?)\}\}",          # {{필드명}}
    r"\{([가-힣a-zA-Z_]+)\}",  # {필드명}
    r"(O{3,})",                 # OOO
    r"(○{3,})",                # ○○○
    r"(_{3,})",                 # ___
]


def detect_placeholders(filepath: Path) -> list[dict]:
    """엑셀 양식에서 플레이스홀더를 탐지한다.

    Returns:
        [
            {
                "placeholder": "{{학과명}}",
                "field_name": "학과명",
                "sheet": "Sheet1",
                "row": 3,
                "col": 2,
                "cell_ref": "B3",
                "context": "학과명: {{학과명}}",
            },
            ...
        ]
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=False)
    results: list[dict] = []
    seen: set[str] = set()

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_str = str(cell.value)

                for pattern in PLACEHOLDER_PATTERNS:
                    for match in re.finditer(pattern, cell_str):
                        placeholder = match.group(0)
                        if placeholder in seen:
                            continue
                        seen.add(placeholder)

                        field_name = match.group(1) if match.lastindex else placeholder
                        col_letter = openpyxl.utils.get_column_letter(cell.column)
                        results.append({
                            "placeholder": placeholder,
                            "field_name": field_name,
                            "sheet": ws.title,
                            "row": cell.row,
                            "col": cell.column,
                            "cell_ref": f"{col_letter}{cell.row}",
                            "context": cell_str.strip()[:80],
                        })

    wb.close()
    return results


def fill_template(
    template_path: Path,
    field_values: dict[str, str],
    output_path: Path | None = None,
    progress_cb: Callable[[int, int, str], None] | None = None,
) -> dict:
    """양식 파일의 플레이스홀더를 값으로 채워넣는다.

    Args:
        template_path: 양식 xlsx 파일 경로.
        field_values: {플레이스홀더_또는_필드명: 채울_값} 딕셔너리.
        output_path: 저장 경로. None이면 자동 생성.
        progress_cb: (현재, 전체, 필드명) 콜백.

    Returns:
        {"output_path": str, "filled_fields": int, "unfilled": list[str]}
    """
    if not template_path.exists():
        raise FileNotFoundError(f"양식 파일을 찾을 수 없습니다: {template_path}")

    if output_path is None:
        output_path = template_path.parent / f"{template_path.stem}_작성완료{template_path.suffix}"

    shutil.copy2(template_path, output_path)

    # 플레이스홀더 탐지
    placeholders = detect_placeholders(template_path)

    # 필드명 → 플레이스홀더 매핑
    field_to_placeholder: dict[str, list[dict]] = {}
    for p in placeholders:
        field_to_placeholder.setdefault(p["field_name"], []).append(p)
        field_to_placeholder.setdefault(p["placeholder"], []).append(p)

    wb = openpyxl.load_workbook(output_path)
    filled_count = 0
    filled_placeholders: set[str] = set()

    total = len(field_values)
    for idx, (key, value) in enumerate(field_values.items()):
        if progress_cb:
            progress_cb(idx + 1, total, key)

        # 방법 1: 직접 매핑된 플레이스홀더 찾기
        targets = field_to_placeholder.get(key, [])
        if targets:
            for target in targets:
                ws = wb[target["sheet"]]
                cell = ws.cell(row=target["row"], column=target["col"])
                # 셀 값에서 플레이스홀더만 교체 (다른 텍스트 보존)
                cell_str = str(cell.value) if cell.value else ""
                cell.value = cell_str.replace(target["placeholder"], str(value))
                filled_placeholders.add(target["placeholder"])
                filled_count += 1
        else:
            # 방법 2: 모든 셀에서 텍스트 직접 치환
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None and key in str(cell.value):
                            cell.value = str(cell.value).replace(key, str(value))
                            filled_count += 1

    wb.save(output_path)
    wb.close()

    # 채우지 못한 플레이스홀더
    unfilled = [p["placeholder"] for p in placeholders if p["placeholder"] not in filled_placeholders]

    return {
        "output_path": str(output_path),
        "filled_fields": filled_count,
        "unfilled": unfilled,
    }


def get_template_summary(filepath: Path) -> dict:
    """양식 파일의 요약 정보를 반환한다.

    Returns:
        {
            "filepath": str,
            "placeholders": [...],
            "total_fields": int,
            "field_names": [str, ...],
            "sheets_with_fields": [str, ...],
        }
    """
    placeholders = detect_placeholders(filepath)
    sheets = list({p["sheet"] for p in placeholders})
    return {
        "filepath": str(filepath),
        "placeholders": placeholders,
        "total_fields": len(placeholders),
        "field_names": [p["field_name"] for p in placeholders],
        "sheets_with_fields": sheets,
    }
