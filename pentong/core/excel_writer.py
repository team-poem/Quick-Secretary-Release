"""엑셀 파일 쓰기 — openpyxl 기반.

셀/범위 쓰기, 행 추가, 서식 보존.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill


def write_cell(
    filepath: Path,
    row: int,
    col: int,
    value,
    sheet_name: str | None = None,
    output_path: Path | None = None,
) -> dict:
    """특정 셀에 값을 쓴다.

    Args:
        filepath: xlsx 파일 경로.
        row: 행 번호 (1부터).
        col: 열 번호 (1부터).
        value: 쓸 값.
        sheet_name: 시트 이름. None이면 활성 시트.
        output_path: 저장 경로. None이면 원본에 덮어쓰기.

    Returns:
        {"output_path": str, "cell": "B3"}
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    if output_path is None:
        output_path = filepath

    if output_path.resolve() != filepath.resolve():
        shutil.copy2(filepath, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    ws.cell(row=row, column=col, value=value)
    wb.save(output_path)
    wb.close()

    col_letter = openpyxl.utils.get_column_letter(col)
    return {"output_path": str(output_path), "cell": f"{col_letter}{row}"}


def write_range(
    filepath: Path,
    data: list[list],
    start_row: int = 1,
    start_col: int = 1,
    sheet_name: str | None = None,
    output_path: Path | None = None,
) -> dict:
    """2차원 데이터를 특정 범위에 쓴다.

    Args:
        filepath: xlsx 파일 경로.
        data: 2차원 리스트 [[값, 값, ...], ...].
        start_row: 시작 행 (1부터).
        start_col: 시작 열 (1부터).
        sheet_name: 시트 이름.
        output_path: 저장 경로.

    Returns:
        {"output_path": str, "rows_written": int, "cols_written": int}
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")
    if not data:
        raise ValueError("쓸 데이터가 비어있습니다.")

    if output_path is None:
        output_path = filepath

    if output_path.resolve() != filepath.resolve():
        shutil.copy2(filepath, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    max_cols = 0
    for r_idx, row_data in enumerate(data):
        for c_idx, val in enumerate(row_data):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
        max_cols = max(max_cols, len(row_data))

    wb.save(output_path)
    wb.close()

    return {
        "output_path": str(output_path),
        "rows_written": len(data),
        "cols_written": max_cols,
    }


def append_rows(
    filepath: Path,
    rows: list[list],
    sheet_name: str | None = None,
    output_path: Path | None = None,
) -> dict:
    """기존 데이터 아래에 행을 추가한다.

    Args:
        filepath: xlsx 파일 경로.
        rows: 추가할 행 데이터 [[값, ...], ...].
        sheet_name: 시트 이름.
        output_path: 저장 경로.

    Returns:
        {"output_path": str, "appended_rows": int, "start_row": int}
    """
    if not filepath.exists():
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")
    if not rows:
        raise ValueError("추가할 행이 비어있습니다.")

    if output_path is None:
        output_path = filepath

    if output_path.resolve() != filepath.resolve():
        shutil.copy2(filepath, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    start_row = (ws.max_row or 0) + 1

    for row_data in rows:
        ws.append(row_data)

    wb.save(output_path)
    wb.close()

    return {
        "output_path": str(output_path),
        "appended_rows": len(rows),
        "start_row": start_row,
    }


def write_dicts(
    filepath: Path,
    data: list[dict],
    sheet_name: str | None = None,
    output_path: Path | None = None,
    write_header: bool = True,
) -> dict:
    """딕셔너리 리스트를 엑셀에 쓴다. 새 시트를 생성하거나 기존 시트에 덮어쓴다.

    Args:
        filepath: xlsx 파일 경로 (없으면 새로 생성).
        data: [{컬럼명: 값}, ...] 리스트.
        sheet_name: 시트 이름.
        output_path: 저장 경로.
        write_header: 헤더(컬럼명) 행 포함 여부.

    Returns:
        {"output_path": str, "rows_written": int, "columns": [str, ...]}
    """
    if not data:
        raise ValueError("쓸 데이터가 비어있습니다.")

    if output_path is None:
        output_path = filepath

    # 컬럼 순서: 첫 번째 dict의 키 순서
    columns = list(data[0].keys())

    if filepath.exists():
        if output_path.resolve() != filepath.resolve():
            shutil.copy2(filepath, output_path)
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()

    if sheet_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 기존 내용 지우기
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        ws = wb.active

    current_row = 1

    if write_header:
        for c_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
        current_row += 1

    for row_dict in data:
        for c_idx, col_name in enumerate(columns, 1):
            ws.cell(row=current_row, column=c_idx, value=row_dict.get(col_name))
        current_row += 1

    wb.save(output_path)
    wb.close()

    return {
        "output_path": str(output_path),
        "rows_written": len(data),
        "columns": columns,
    }


def create_new_workbook(
    output_path: Path,
    sheet_name: str = "Sheet1",
    headers: list[str] | None = None,
) -> dict:
    """새 엑셀 파일을 생성한다.

    Args:
        output_path: 생성할 파일 경로.
        sheet_name: 첫 시트 이름.
        headers: 헤더 행. None이면 빈 시트.

    Returns:
        {"output_path": str, "sheet_name": str}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers:
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return {"output_path": str(output_path), "sheet_name": sheet_name}
