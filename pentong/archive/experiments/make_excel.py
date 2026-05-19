import win32com.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

hwp = win32com.client.Dispatch('HWPFrame.HwpObject')
hwp.RegisterModule('FilePathCheckDLL', 'SecurityModule')
hwp.XHwpWindows.Item(0).Visible = False
path = r'C:\Users\user\Desktop\pentong-20260416T024438Z-3-001\pentong\2025교육혁신처_20260409.hwp'
hwp.Open(path, 'HWP', 'forceopen:true')
text = hwp.GetTextFile('TEXT', '')
hwp.Quit()

lines = text.split('\n')
lc = [l.rstrip('\r') for l in lines]
print(f'총 라인 수: {len(lc)}')

byo5 = [
    ("AI공학", "소프트웨어학과, 컴퓨터공학과, 인공지능응용학과, 정보보안학과, 게임학과, 컴퓨터공학부, 디지털콘텐츠학부", "인공지능공학사"),
    ("빅데이터", "소프트웨어학과, 컴퓨터공학과, 인공지능응용학과, 정보보안학과, 게임학과, 컴퓨터공학부, 디지털콘텐츠학부", "빅데이터학사"),
    ("AI콘텐츠", "영상애니메이션학과, 게임학과, 소프트웨어학과, 빅데이터전공", "인공지능콘텐츠학사"),
    ("사이버경찰보안", "경찰행정학과, 정보보안학과", "사이버경찰보안학사"),
    ("인공지능로봇", "기계공학과, 소프트웨어학과, 인공지능응용학과, AI공학전공", "인공지능공학사"),
    ("AI경험디자인", "디지털미디어디자인전공, 제품인터렉션디자인전공, 소프트웨어학과, 컴퓨터공학과", "디자인공학사"),
    ("XR콘텐츠", "게임학과, 영상애니메이션학과, 방송영상학과, 인공지능응용학과", "디지털콘텐츠학사"),
    ("영어SW", "영어학과, 소프트웨어학과, AI공학전공, 빅데이터전공", "SW융합학사"),
    ("일본어SW", "일본어학과, 소프트웨어학과, AI공학전공, 빅데이터전공", "SW융합학사"),
    ("중국어SW", "중국어학과, 소프트웨어학과, AI공학전공, 빅데이터전공", "SW융합학사"),
    ("글로벌 K-컬처 영화/드라마", "영상애니메이션학과, 영화과", "융합학사"),
    ("글로벌 K-컬처 웹툰애니메이션", "영상애니메이션학과, 영화과", "융합학사"),
    ("글로벌 K-컬처 음악/공연", "영상애니메이션학과, 영화과", "융합학사"),
    ("글로벌 K-컬처 문학/웹소설", "영상애니메이션학과, 영화과", "융합학사"),
    ("스마트 국제물류", "국제통상학과, 국제물류학과", "스마트 국제물류 경영학사"),
    ("스마트 항만물류", "국제통상학과, 국제물류학과", "스마트 항만물류 공학사"),
    ("해양미래산업", "국제물류학과", "해양미래산업 공학사"),
    ("스마트 해양모빌리티", "컴퓨터공학과, 소프트웨어학과, 인공지능응용학과, 정보보안학과", "공학사"),
    ("미래모빌리티", "컴퓨터공학과, 정보보안학과, 전기전자공학전공", "공학사"),
    ("AI혁신", "컴퓨터공학과, 소프트웨어학과", "공학사"),
    ("클린에너지", "수소에너지전공, 이차전지전공", "클린에너지 공학사"),
    ("평생교육컨설팅", "사회복지상담학과, 스포츠레저산업학과", "평생교육학사"),
    ("ESG디자인", "시각디자인전공, 환경디자인전공, 패션디자인학과, 광고홍보학과, 제품인터랙션디자인전공, 디지털미디어디자인전공, 영상애니메이션학과", "디자인학사"),
    ("전기배터리", "전기전자공학과, 화장품학과, 기계공학과, 컴퓨터공학과, ICT융합공학과, 인공지능응용학과, 스마트모빌리티학부", "공학사"),
    ("문화콘텐츠 크리에이터", "방송영상학과, 경영학전공, 광고홍보학과, 영상애니메이션학과, 웹툰학과, 디자인학부", "문화콘텐츠학사"),
    ("수소에너지", "스마트모빌리티학부", "클린에너지 공학사"),
    ("전력반도체", "스마트모빌리티학부", "전력반도체 공학사"),
    ("첨단콘텐츠", "방송영상학과, 게임학과, 영상애니메이션학과, 웹툰학과, 영화과, 뮤지컬엔터테인먼트과, 연기과, 소프트웨어학과, 컴퓨터공학과", "예술공학사"),
    ("융합디자인", "시각디자인전공, 디지털미디어디자인전공, 제품인터랙션디자인전공, 환경디자인전공, 패션디자인학과, 광고홍보학과, 건축학과", "융합디자인학사"),
    ("휴먼메타케어", "운동처방학과, 식품영양학과, 시니어운동처방학과", "휴먼메타케어학사"),
    ("헤리티지콘텐츠", "방송영상학과, 게임학과, 관광경영컨벤션학과, 시각디자인전공, 웹툰학과, 영상애니메이션학과, 디지털미디어디자인전공, 제품인터랙션디자인전공", "헤리티지콘텐츠학사"),
    ("Applied Artificial Intelligence", "Computer Science학과, Game Development학과", "인공지능공학사"),
    ("Digital Marketing", "Global Business Administration학과, Digital Design학과", "디지털마케팅학사"),
    ("자기설계", "소속전공을 포함한 3개 학과 이상의 전공교과목으로 융합전공 설계", "문학사, 이학사, 공학사, 예술학사 중 택1"),
]

courses = []
i = 18621
while i < min(len(lc), 19200):
    g = lc[i].strip()
    if g not in ['전공필수', '전공선택']:
        i += 1
        continue
    if i + 4 < len(lc):
        code = lc[i+1].strip()
        nm   = lc[i+2].strip()
        grd  = lc[i+3].strip()
        cr   = lc[i+4].strip()
        dp   = lc[i+5].strip() if i+5 < len(lc) else ''
        if code.isdigit() and len(code) == 6:
            courses.append((g, code, nm, grd, cr, dp))
            i += 6
        else:
            i += 1
    else:
        i += 1

print(f'[별표5] 항목: {len(byo5)}, AI콘텐츠 교과목: {len(courses)}')

thin = Side(style='thin')
BOR  = Border(left=thin, right=thin, top=thin, bottom=thin)
HF   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
AF   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RF   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
AI_KW = ['AI', '인공지능', 'Artificial Intelligence']

def wtitle(ws, t, cols):
    ec = chr(64+cols)
    ws.merge_cells(f'A1:{ec}1')
    c = ws['A1']
    c.value = t
    c.font  = Font(bold=True, size=13)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

def wheader(ws, hdrs, row=2):
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = HF
        c.alignment = CTR
        c.border = BOR
    ws.row_dimensions[row].height = 22

def wc(ws, row, col, val, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.border = BOR
    c.alignment = align or LFT
    if fill:
        c.fill = fill

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = '융합연계전공_별표5'
wtitle(ws1, '[별표 5] 융합연계전공 목록 (2025교육혁신처_20260409)', 3)
wheader(ws1, ['(융합)연계전공', '참여학부(전공)', '학위'])
for r, (nm, dept, deg) in enumerate(byo5, start=3):
    f = AF if any(k in nm for k in AI_KW) else None
    wc(ws1, r, 1, nm,   fill=f, align=CTR)
    wc(ws1, r, 2, dept, fill=f)
    wc(ws1, r, 3, deg,  fill=f, align=CTR)
    ws1.row_dimensions[r].height = 30
ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 72
ws1.column_dimensions['C'].width = 27
nr = len(byo5)+4
ws1.cell(row=nr, column=1, value='※ 파란색 배경: AI/인공지능 관련 융합연계전공').font = Font(italic=True, color='1F4E79', size=10)

ws2 = wb.create_sheet('AI공학_융합연계전공')
wtitle(ws2, 'AI·인공지능 관련 융합연계전공 목록', 3)
wheader(ws2, ['(융합)연계전공', '참여학부(전공)', '학위'])
ai_rows = [(n,d,dg) for n,d,dg in byo5 if any(k in n for k in AI_KW)]
for r, (nm, dept, deg) in enumerate(ai_rows, start=3):
    wc(ws2, r, 1, nm,   fill=AF, align=CTR)
    wc(ws2, r, 2, dept, fill=AF)
    wc(ws2, r, 3, deg,  fill=AF, align=CTR)
    ws2.row_dimensions[r].height = 35
ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 72
ws2.column_dimensions['C'].width = 27
nr2 = len(ai_rows)+4
ws2.merge_cells(f'A{nr2}:C{nr2}')
ws2.cell(row=nr2, column=1, value='※ "가. AI공학 융합연계전공" 본문은 원본 문서에 미기재 상태(수정예정)입니다.').font = Font(italic=True, color='FF0000', size=10)

ws3 = wb.create_sheet('AI콘텐츠_교과목')
wtitle(ws3, 'AI콘텐츠 융합연계전공 주요 교과목', 6)
wheader(ws3, ['이수구분', '교과목코드', '교과목명', '학년/학기', '학점', '개설전공'])
for r, (g, code, nm, grd, cr, dp) in enumerate(courses, start=3):
    f = RF if g=='전공필수' else None
    wc(ws3, r, 1, g,    fill=f, align=CTR)
    wc(ws3, r, 2, code, fill=f, align=CTR)
    wc(ws3, r, 3, nm,   fill=f)
    wc(ws3, r, 4, grd,  fill=f, align=CTR)
    wc(ws3, r, 5, cr,   fill=f, align=CTR)
    wc(ws3, r, 6, dp,   fill=f, align=CTR)
    ws3.row_dimensions[r].height = 20
ws3.column_dimensions['A'].width = 13
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 38
ws3.column_dimensions['D'].width = 13
ws3.column_dimensions['E'].width = 8
ws3.column_dimensions['F'].width = 15
nr3 = len(courses)+4
ws3.cell(row=nr3, column=1, value='※ 노란색 배경: 전공필수 과목').font = Font(italic=True, size=10)

out = r'C:\Users\user\Desktop\pentong-20260416T024438Z-3-001\pentong\융합연계전공_AI공학_2025교육혁신처.xlsx'
wb.save(out)
print(f'Excel 저장 완료: {out}')
