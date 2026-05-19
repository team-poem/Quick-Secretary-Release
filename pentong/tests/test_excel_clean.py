"""엑셀 공백 행 제거 기능 테스트."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from core.excel_clean import count_blank_rows, remove_blank_rows


def _make_wb(tmp_path: Path, rows: list[list]) -> Path:
    """주어진 행 데이터로 xlsx 파일을 생성한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def _read_values(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows()]


# ---------------------------------------------------------------------------
# count_blank_rows
# ---------------------------------------------------------------------------

def test_count_blank_rows_none(tmp_path: Path) -> None:
    path = _make_wb(tmp_path, [["이름", "나이"], ["홍길동", 30]])
    assert count_blank_rows(path) == 0


def test_count_blank_rows_with_blanks(tmp_path: Path) -> None:
    path = _make_wb(tmp_path, [["이름"], [None], ["", ""], ["홍길동"]])
    assert count_blank_rows(path) == 2


def test_count_blank_rows_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        count_blank_rows(tmp_path / "없는파일.xlsx")


# ---------------------------------------------------------------------------
# remove_blank_rows
# ---------------------------------------------------------------------------

def test_remove_blank_rows_no_blanks(tmp_path: Path) -> None:
    rows = [["이름", "나이"], ["홍길동", 30], ["김철수", 25]]
    path = _make_wb(tmp_path, rows)
    out = tmp_path / "out.xlsx"
    removed = remove_blank_rows(path, out)
    assert removed == 0
    assert _read_values(out) == rows


def test_remove_blank_rows_consecutive(tmp_path: Path) -> None:
    rows = [["이름"], [None], [None], ["홍길동"], ["", " "], ["김철수"]]
    path = _make_wb(tmp_path, rows)
    out = tmp_path / "out.xlsx"
    removed = remove_blank_rows(path, out)
    assert removed == 3
    result = _read_values(out)
    assert result == [["이름"], ["홍길동"], ["김철수"]]


def test_remove_blank_rows_overwrite(tmp_path: Path) -> None:
    rows = [["A"], [None], ["B"]]
    path = _make_wb(tmp_path, rows)
    removed = remove_blank_rows(path, path)
    assert removed == 1
    assert _read_values(path) == [["A"], ["B"]]


def test_remove_blank_rows_styled_blank(tmp_path: Path) -> None:
    """서식이 있지만 값이 없는 행은 공백으로 처리한다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["헤더"])
    ws.append([None])  # blank row
    ws.append(["데이터"])
    # 공백 행에 배경색 서식 적용
    from openpyxl.styles import PatternFill
    ws.cell(row=2, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    path = tmp_path / "styled.xlsx"
    wb.save(path)

    out = tmp_path / "out.xlsx"
    removed = remove_blank_rows(path, out)
    assert removed == 1
    result = _read_values(out)
    assert result == [["헤더"], ["데이터"]]


def test_remove_blank_rows_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        remove_blank_rows(tmp_path / "없는파일.xlsx", tmp_path / "out.xlsx")
