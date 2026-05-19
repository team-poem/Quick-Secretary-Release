"""core/excel_merge.py 단위 테스트."""

import pytest
from pathlib import Path

import openpyxl

from core.excel_merge import merge_excel_files


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _make_xlsx(path: Path, rows: list[list]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# 테스트
# ---------------------------------------------------------------------------

def test_merge_empty_list_raises(tmp_path):
    out = tmp_path / "out.xlsx"
    with pytest.raises(ValueError, match="하나 이상"):
        merge_excel_files([], out)


def test_merge_missing_file_raises(tmp_path):
    out = tmp_path / "out.xlsx"
    with pytest.raises(FileNotFoundError):
        merge_excel_files([tmp_path / "ghost.xlsx"], out)


def test_merge_sheet_mode_creates_separate_sheets(tmp_path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_xlsx(a, [["이름", "나이"], ["홍길동", 30]])
    _make_xlsx(b, [["이름", "나이"], ["김철수", 25]])

    out = tmp_path / "merged.xlsx"
    merge_excel_files([a, b], out, mode="sheet")

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == "a"
    assert wb.sheetnames[1] == "b"
    assert wb[wb.sheetnames[0]].cell(2, 1).value == "홍길동"
    assert wb[wb.sheetnames[1]].cell(2, 1).value == "김철수"


def test_merge_row_mode_concatenates_rows(tmp_path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_xlsx(a, [["이름", "나이"], ["홍길동", 30]])
    _make_xlsx(b, [["이름", "나이"], ["김철수", 25]])

    out = tmp_path / "merged.xlsx"
    merge_excel_files([a, b], out, mode="row")

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # 헤더 1행 + 데이터 2행 = 3행 (두 번째 파일 헤더 제거)
    assert ws.max_row == 3
    assert ws.cell(1, 1).value == "이름"
    assert ws.cell(2, 1).value == "홍길동"
    assert ws.cell(3, 1).value == "김철수"


def test_merge_different_columns_row_mode(tmp_path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_xlsx(a, [["A", "B"], [1, 2]])
    _make_xlsx(b, [["A", "B", "C"], [3, 4, 5]])

    out = tmp_path / "merged.xlsx"
    # 열 구조가 다르더라도 오류 없이 완료되어야 함
    merge_excel_files([a, b], out, mode="row")

    wb = openpyxl.load_workbook(out)
    assert wb.active.max_row == 3


def test_merge_progress_callback_called(tmp_path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _make_xlsx(a, [["x"]])
    _make_xlsx(b, [["x"]])

    calls: list[tuple[int, int, str]] = []
    merge_excel_files([a, b], tmp_path / "out.xlsx", mode="sheet",
                      progress_cb=lambda c, t, n: calls.append((c, t, n)))

    assert len(calls) == 2
    assert calls[0] == (1, 2, "a.xlsx")
    assert calls[1] == (2, 2, "b.xlsx")


def test_merge_sheet_mode_duplicate_names(tmp_path):
    """같은 이름의 파일이 여러 개일 때 시트 이름 중복을 피해야 한다."""
    subdir1 = tmp_path / "d1"
    subdir2 = tmp_path / "d2"
    subdir1.mkdir()
    subdir2.mkdir()
    a = subdir1 / "report.xlsx"
    b = subdir2 / "report.xlsx"
    _make_xlsx(a, [["v", 1]])
    _make_xlsx(b, [["v", 2]])

    out = tmp_path / "merged.xlsx"
    merge_excel_files([a, b], out, mode="sheet")

    wb = openpyxl.load_workbook(out)
    assert len(set(wb.sheetnames)) == 2
