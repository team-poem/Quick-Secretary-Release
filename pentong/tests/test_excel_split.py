"""core/excel_split.py 단위 테스트."""

from __future__ import annotations

import pytest
from pathlib import Path

import openpyxl

from core.excel_split import split_excel_file


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------

def _make_xlsx(path: Path, rows: list[list], sheet_names: list[str] | None = None) -> None:
    wb = openpyxl.Workbook()
    sheets = sheet_names or ["Sheet"]
    wb.active.title = sheets[0]
    ws = wb.active
    for row in rows:
        ws.append(row)
    for name in sheets[1:]:
        wb.create_sheet(title=name)
    wb.save(path)


def _make_multisheet_xlsx(path: Path, sheet_data: dict[str, list[list]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheet_data.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# 에러 케이스
# ---------------------------------------------------------------------------

def test_split_missing_file_raises(tmp_path):
    with pytest.raises(FileNotFoundError):
        split_excel_file(tmp_path / "ghost.xlsx", tmp_path / "out")


def test_split_invalid_rows_raises(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_xlsx(f, [["x"]])
    with pytest.raises(ValueError, match="rows_per_file"):
        split_excel_file(f, tmp_path / "out", mode="rows", rows_per_file=0)


def test_split_invalid_column_raises(tmp_path):
    f = tmp_path / "a.xlsx"
    _make_xlsx(f, [["x"]])
    with pytest.raises(ValueError, match="column_index"):
        split_excel_file(f, tmp_path / "out", mode="column", column_index=0)


# ---------------------------------------------------------------------------
# 시트별 분할
# ---------------------------------------------------------------------------

def test_split_by_sheet_single_sheet(tmp_path):
    src = tmp_path / "data.xlsx"
    _make_xlsx(src, [["이름", "나이"], ["홍길동", 30]])

    results = split_excel_file(src, tmp_path / "out", mode="sheet")

    assert len(results) == 1
    wb = openpyxl.load_workbook(results[0])
    assert wb.active.cell(2, 1).value == "홍길동"


def test_split_by_sheet_multiple_sheets(tmp_path):
    src = tmp_path / "multi.xlsx"
    _make_multisheet_xlsx(src, {
        "영업": [["항목", "금액"], ["A", 100]],
        "회계": [["항목", "금액"], ["B", 200]],
    })

    out_dir = tmp_path / "out"
    results = split_excel_file(src, out_dir, mode="sheet")

    assert len(results) == 2
    names = {p.stem for p in results}
    assert "multi_영업" in names
    assert "multi_회계" in names

    for p in results:
        wb = openpyxl.load_workbook(p)
        assert len(wb.sheetnames) == 1


def test_split_by_sheet_creates_output_dir(tmp_path):
    src = tmp_path / "a.xlsx"
    _make_xlsx(src, [["x"]])
    out_dir = tmp_path / "new_dir" / "nested"

    results = split_excel_file(src, out_dir, mode="sheet")

    assert out_dir.exists()
    assert len(results) == 1


# ---------------------------------------------------------------------------
# 행 수 기준 분할
# ---------------------------------------------------------------------------

def test_split_by_rows_basic(tmp_path):
    src = tmp_path / "rows.xlsx"
    _make_xlsx(src, [["이름"]] + [[f"행{i}"] for i in range(10)])

    out_dir = tmp_path / "out"
    results = split_excel_file(src, out_dir, mode="rows", rows_per_file=3)

    # 10개 데이터 행 / 3 = 4개 파일 (3+3+3+1)
    assert len(results) == 4
    for p in results:
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        # 헤더가 모든 파일에 존재
        assert ws.cell(1, 1).value == "이름"


def test_split_by_rows_header_not_duplicated(tmp_path):
    src = tmp_path / "rows.xlsx"
    _make_xlsx(src, [["헤더"], ["데이터1"], ["데이터2"]])

    results = split_excel_file(src, tmp_path / "out", mode="rows", rows_per_file=1)

    assert len(results) == 2
    for p in results:
        wb = openpyxl.load_workbook(p)
        # 각 파일은 헤더 1행 + 데이터 1행 = 2행
        assert wb.active.max_row == 2


def test_split_by_rows_progress_callback(tmp_path):
    src = tmp_path / "rows.xlsx"
    _make_xlsx(src, [["h"]] + [[i] for i in range(6)])

    calls: list[tuple[int, int, str]] = []
    split_excel_file(
        src, tmp_path / "out", mode="rows", rows_per_file=3,
        progress_cb=lambda c, t, n: calls.append((c, t, n))
    )

    assert len(calls) == 2
    assert calls[0][0] == 1
    assert calls[1][0] == 2


# ---------------------------------------------------------------------------
# 열 값 기준 분할
# ---------------------------------------------------------------------------

def test_split_by_column_basic(tmp_path):
    src = tmp_path / "col.xlsx"
    _make_xlsx(src, [
        ["부서", "이름"],
        ["영업", "홍길동"],
        ["영업", "이순신"],
        ["회계", "김철수"],
    ])

    results = split_excel_file(src, tmp_path / "out", mode="column", column_index=1)

    assert len(results) == 2
    stems = {p.stem for p in results}
    assert "col_영업" in stems
    assert "col_회계" in stems

    for p in results:
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        assert ws.cell(1, 1).value == "부서"  # 헤더 확인


def test_split_by_column_empty_value(tmp_path):
    """빈 셀 값은 __빈값__ 그룹으로 처리되어야 한다."""
    src = tmp_path / "col.xlsx"
    _make_xlsx(src, [
        ["구분", "값"],
        ["A", 1],
        [None, 2],
        ["A", 3],
    ])

    results = split_excel_file(src, tmp_path / "out", mode="column", column_index=1)

    assert len(results) == 2
    stems = {p.stem for p in results}
    assert "col_A" in stems
    assert any("빈값" in s for s in stems)


def test_split_by_column_row_counts(tmp_path):
    src = tmp_path / "col.xlsx"
    _make_xlsx(src, [
        ["그룹"],
        ["X"],
        ["X"],
        ["Y"],
    ])

    results = split_excel_file(src, tmp_path / "out", mode="column", column_index=1)

    by_stem = {p.stem: p for p in results}
    wb_x = openpyxl.load_workbook(by_stem["col_X"])
    # 헤더 1행 + X 데이터 2행
    assert wb_x.active.max_row == 3

    wb_y = openpyxl.load_workbook(by_stem["col_Y"])
    # 헤더 1행 + Y 데이터 1행
    assert wb_y.active.max_row == 2
